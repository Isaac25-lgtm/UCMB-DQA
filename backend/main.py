from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
import csv
import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

from database import engine, get_db, Base
from models import Facility, Indicator, DqaSession, DqaLine

# Optional AI integration for report summaries
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

def generate_ai_summary(stats, sessions, lines):
    """Generate AI-powered executive summary using Google Gemini API"""
    if not GEMINI_AVAILABLE:
        return None
    
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        return None
    
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-pro')
        
        # Calculate additional statistics
        completion_rate = round((stats['assessed_facilities'] / stats['total_facilities'] * 100) if stats['total_facilities'] > 0 else 0)
        
        # Count deviations
        red_count = 0
        amber_count = 0
        green_count = 0
        for line in lines:
            if line.dev_dhis2_vs_reg is not None:
                abs_dev = abs(line.dev_dhis2_vs_reg)
                if abs_dev > 0.10:
                    red_count += 1
                elif abs_dev > 0.05:
                    amber_count += 1
                else:
                    green_count += 1
        
        # Build team performance summary
        team_summary = "\n".join([f"- {t['team']}: {t['facilities_assessed']} facilities assessed" for t in stats['team_progress']])
        
        prompt = f"""Generate a professional executive summary for a Data Quality Assurance (DQA) assessment report for Maternal and Newborn Health (MNH) indicators.

Assessment Statistics:
- Total Facilities: {stats['total_facilities']}
- Facilities Assessed: {stats['assessed_facilities']}
- Completion Rate: {completion_rate}%
- Total Assessment Sessions: {stats['total_sessions']}

Data Quality Findings:
- High Quality (≤5% deviation): {green_count} indicators
- Moderate Quality (5-10% deviation): {amber_count} indicators
- Low Quality (>10% deviation): {red_count} indicators

Team Performance:
{team_summary}

Please provide a concise 3-paragraph executive summary that:
1. Highlights key achievements and overall progress
2. Identifies areas requiring attention based on data quality findings
3. Provides actionable recommendations for improvement

Write in a professional, clear, and concise style suitable for healthcare management."""
        
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        print(f"Gemini API error: {e}")
        return None

def generate_ai_team_insights(stats, sessions):
    """Generate AI-powered insights for team performance"""
    if not GEMINI_AVAILABLE:
        return None
    
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        return None
    
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-pro')
        
        # Calculate sessions per team
        team_sessions = {}
        for session in sessions:
            team = session.get('team', 'Unknown')
            if team not in team_sessions:
                team_sessions[team] = 0
            team_sessions[team] += 1
        
        # Build detailed team data
        team_data = []
        for team_info in stats['team_progress']:
            team_name = team_info['team']
            team_data.append(f"{team_name}: {team_info['facilities_assessed']} facilities, {team_sessions.get(team_name, 0)} sessions")
        
        team_details = "\n".join(team_data)
        
        prompt = f"""Analyze team performance data from a DQA assessment and provide insights:

Team Performance Data:
{team_details}

Provide 2-3 bullet points of insights:
- Identify top performing teams
- Highlight teams that may need additional support
- Suggest strategies for improvement

Keep it concise and actionable."""
        
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        print(f"Gemini API error (team insights): {e}")
        return None
from schemas import (
    Facility as FacilitySchema,
    Indicator as IndicatorSchema,
    DqaSessionCreate,
    DqaSessionResponse,
    DqaSessionSummary,
    DqaLineResponse
)
from crud import (
    get_facilities,
    get_indicators,
    get_facility,
    get_indicator_by_code,
    get_sessions,
    get_session,
    create_session,
    delete_session,
    get_dashboard_stats
)

app = FastAPI()

# CORS middleware
# Allow both localhost (development) and Render (production) origins
import os

# Try to load .env file if python-dotenv is available (for local development)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv not installed, use environment variables directly
allowed_origins = [
    "http://localhost:5173",
    "http://localhost:3000",
]

# Add any Render frontend URL from environment variable
frontend_url = os.getenv("FRONTEND_URL")
if frontend_url:
    allowed_origins.append(frontend_url)

# For production, also allow common Render patterns
# This allows any *.onrender.com domain
allowed_origins.append("https://*.onrender.com")

# If no specific frontend URL is set, allow all origins (for easier setup)
# Remove this in production if you want stricter CORS
if not frontend_url and os.getenv("RENDER"):
    allowed_origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins if allowed_origins != ["*"] else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create tables
Base.metadata.create_all(bind=engine)

# Add comments column if it doesn't exist (migration)
def migrate_database():
    """Add new columns to existing database"""
    conn = engine.connect()
    try:
        # Check if comments column exists
        result = conn.execute(text("PRAGMA table_info(dqa_sessions)"))
        columns = [row[1] for row in result]
        if 'comments' not in columns:
            conn.execute(text("ALTER TABLE dqa_sessions ADD COLUMN comments TEXT"))
            conn.commit()
    except Exception as e:
        print(f"Migration note: {e}")
    finally:
        conn.close()

migrate_database()

# Seed data
def seed_data():
    db = next(get_db())
    try:
        # Remove deprecated indicators if they exist in an existing database
        deprecated_codes = ["AN01", "AN02", "AN11", "PN01", "MA22", "MA05b", "MA05c", "MA12"]
        if deprecated_codes:
            # First, delete associated DqaLine records
            deprecated_indicators = db.query(Indicator).filter(Indicator.code.in_(deprecated_codes)).all()
            if deprecated_indicators:
                indicator_ids = [ind.id for ind in deprecated_indicators]
                db.query(DqaLine).filter(DqaLine.indicator_id.in_(indicator_ids)).delete(
                    synchronize_session=False
                )
            # Then delete the indicators
            db.query(Indicator).filter(Indicator.code.in_(deprecated_codes)).delete(
                synchronize_session=False
            )

        # Check if facilities exist
        if db.query(Facility).count() == 0:
            facilities_data = [
                {"name": "Agoro Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Akworo Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Alero Health Centre III", "district": "Nwoya District", "level": "HC III"},
                {"name": "Amuru Lacor Health Centre III", "district": "Amuru District", "level": "HC III"},
                {"name": "Anaka General Hospital", "district": "Nwoya District", "level": "Hospital"},
                {"name": "Atiak Health Centre IV", "district": "Amuru District", "level": "HC IV"},
                {"name": "Awach Health Centre IV", "district": "Gulu District", "level": "HC IV"},
                {"name": "Awich Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Cwero Health Centre III", "district": "Gulu District", "level": "HC III"},
                {"name": "David Fagerlee's Medical Centre", "district": "Agago District", "level": "Medical Centre"},
                {"name": "Dr. Ambrosoli Memorial Hospital Kalongo", "district": "Agago District", "level": "Hospital"},
                {"name": "Kitgum General Hospital", "district": "Kitgum District", "level": "Hospital"},
                {"name": "Kitgum-Matidi Health Centre III", "district": "Kitgum District", "level": "HC III"},
                {"name": "Koch Goma Health Centre III", "district": "Nwoya District", "level": "HC III"},
                {"name": "Labongogali Health Centre III", "district": "Amuru District", "level": "HC III"},
                {"name": "Labworomor Health Centre III", "district": "Gulu District", "level": "HC III"},
                {"name": "Lacor Opit Health Centre III", "district": "Omoro District", "level": "HC III"},
                {"name": "Lacor-Pabbo Health Centre III", "district": "Amuru District", "level": "HC III"},
                {"name": "Lalogi Health Centre IV", "district": "Omoro District", "level": "HC IV"},
                {"name": "Lanenober Health Centre III", "district": "Omoro District", "level": "HC III"},
                {"name": "Layibi Techo Health Centre III", "district": "Gulu City", "level": "HC III"},
                {"name": "Lokung Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Madi-Opei Health Centre IV", "district": "Lamwo District", "level": "HC IV"},
                {"name": "Mary Queen Of Peace Health Centre III", "district": "Gulu City", "level": "HC III"},
                {"name": "Minakulu (Bobi) Health Centre II", "district": "Omoro District", "level": "HC II"},
                {"name": "Mucwini Health Centre III", "district": "Kitgum District", "level": "HC III"},
                {"name": "Namokora Health Centre IV", "district": "Kitgum District", "level": "HC IV"},
                {"name": "Odek Health Centre III", "district": "Omoro District", "level": "HC III"},
                {"name": "Ogonyo Health Centre III", "district": "Pader District", "level": "HC III"},
                {"name": "Okidi Health Centre III", "district": "Kitgum District", "level": "HC III"},
                {"name": "Okinga Health Centre III", "district": "Pader District", "level": "HC III"},
                {"name": "Pabbo Health Centre III", "district": "Amuru District", "level": "HC III"},
                {"name": "Pabwo Health Centre III", "district": "Gulu District", "level": "HC III"},
                {"name": "Pader Health Centre III", "district": "Pader District", "level": "HC III"},
                {"name": "Padibe Health Centre IV", "district": "Lamwo District", "level": "HC IV"},
                {"name": "Padibe West Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Paimol Health Centre III", "district": "Agago District", "level": "HC III"},
                {"name": "Pajule Health Centre IV", "district": "Pader District", "level": "HC IV"},
                {"name": "Palabek-Gem Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Palabek-Kal Health Centre IV", "district": "Lamwo District", "level": "HC IV"},
                {"name": "Paloga Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Paluda Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Patongo Health Centre IV", "district": "Agago District", "level": "HC IV"},
                {"name": "Puranga Health Centre III", "district": "Pader District", "level": "HC III"},
                {"name": "Purongo Health Centre III", "district": "Nwoya District", "level": "HC III"},
                {"name": "St. Joseph's Kitgum Hospital", "district": "Kitgum District", "level": "Hospital"},
                {"name": "St. Mary's Hospital Lacor", "district": "Gulu City", "level": "Hospital"},
                {"name": "St. Mauritz Health Centre III", "district": "Gulu City", "level": "HC III"},
                {"name": "St. Peter and Paul Health Centre III", "district": "Lamwo District", "level": "HC III"},
                {"name": "Todora Health Centre III", "district": "Nwoya District", "level": "HC III"},
            ]
            for fac_data in facilities_data:
                facility = Facility(**fac_data)
                db.add(facility)
        
        # Check if indicators exist
        if db.query(Indicator).count() == 0:
            indicators_data = [
                {"code": "MA04", "name": "Total deliveries in unit", "data_source": "Maternity/delivery register"},
                {"code": "MA05a", "name": "Live births <2.5 kg", "data_source": "Maternity register"},
                {"code": "MA13", "name": "Maternal deaths", "data_source": "Maternity register + death register"},
                {"code": "MA14", "name": "Mothers who initiated breastfeeding within 1st hour after delivery", "data_source": "Maternity register"},
                {"code": "MA23", "name": "Babies with birth asphyxia", "data_source": "Maternity/newborn register"},
                {"code": "MA24", "name": "Babies successfully resuscitated", "data_source": "Maternity/newborn register"},
                {"code": "MA08", "name": "LBW babies <2.5 kg initiated on KMC", "data_source": "KMC register / maternity register"},
            ]
            for ind_data in indicators_data:
                indicator = Indicator(**ind_data)
                db.add(indicator)
        
        db.commit()
    finally:
        db.close()

# Run seed on startup
seed_data()

def calculate_deviations(recount_register, figure_105, figure_dhis2):
    """Calculate deviation percentages"""
    dev_dhis2_vs_reg = None
    dev_105_vs_reg = None
    dev_105_vs_dhis2 = None
    
    if recount_register not in (None, 0):
        if figure_dhis2 is not None:
            dev_dhis2_vs_reg = (figure_dhis2 - recount_register) / recount_register
        if figure_105 is not None:
            dev_105_vs_reg = (figure_105 - recount_register) / recount_register
    
    if figure_dhis2 not in (None, 0) and figure_105 is not None:
        dev_105_vs_dhis2 = (figure_105 - figure_dhis2) / figure_dhis2
    
    return dev_dhis2_vs_reg, dev_105_vs_reg, dev_105_vs_dhis2

@app.get("/facilities", response_model=List[FacilitySchema])
def list_facilities(db: Session = Depends(get_db)):
    return get_facilities(db)

@app.get("/indicators", response_model=List[IndicatorSchema])
def list_indicators(db: Session = Depends(get_db)):
    return get_indicators(db)

@app.get("/teams")
def get_teams():
    """Get list of teams with their members"""
    return {
        "Team A": [
            "Biostat Gulu District",
            "Biostat Pader",
            "Biostat Gulu City",
            "Biostat Kitgum",
        ],
        "Team B": [
            "Biostat Nwoya",
            "Biostat Omoro",
            "Biostat Lamwo",
            "Biostat Gulu RRH",
        ],
        "Team C": [
            "Biostat Amuru",
            "Biostat Lacor",
            "Biostat Agago",
            "Abalo Jenda (UCMB Staff)",
        ],
    }

@app.get("/sessions", response_model=List[DqaSessionSummary])
def list_sessions(db: Session = Depends(get_db)):
    sessions = get_sessions(db)
    return sessions

@app.get("/dashboard/stats")
def get_dashboard_stats_endpoint(db: Session = Depends(get_db)):
    """Get dashboard statistics for graphs"""
    return get_dashboard_stats(db)

@app.get("/sessions/{session_id}", response_model=DqaSessionResponse)
def get_session_detail(session_id: int, db: Session = Depends(get_db)):
    session = get_session(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return session

@app.delete("/sessions/{session_id}")
def delete_dqa_session(session_id: int, db: Session = Depends(get_db)):
    """Delete a DQA session"""
    result = delete_session(db, session_id)
    if not result:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"message": "Session deleted successfully"}

@app.post("/sessions", response_model=DqaSessionResponse)
def create_dqa_session(session_data: DqaSessionCreate, db: Session = Depends(get_db)):
    # Verify facility exists
    facility = get_facility(db, session_data.facility_id)
    if not facility:
        raise HTTPException(status_code=404, detail="Facility not found")
    
    # Prepare session data
    session_dict = {
        "facility_id": session_data.facility_id,
        "period": session_data.period,
        "team": session_data.team,
        "comments": session_data.comments
    }
    
    # Prepare lines with calculated deviations
    lines_list = []
    for line in session_data.lines:
        # Verify indicator exists
        indicator = db.query(Indicator).filter(Indicator.id == line.indicator_id).first()
        if not indicator:
            raise HTTPException(status_code=404, detail=f"Indicator {line.indicator_id} not found")
        
        dev_dhis2_vs_reg, dev_105_vs_reg, dev_105_vs_dhis2 = calculate_deviations(
            line.recount_register, line.figure_105, line.figure_dhis2
        )
        
        lines_list.append({
            "indicator_id": line.indicator_id,
            "recount_register": line.recount_register,
            "figure_105": line.figure_105,
            "figure_dhis2": line.figure_dhis2,
            "dev_dhis2_vs_reg": dev_dhis2_vs_reg,
            "dev_105_vs_reg": dev_105_vs_reg,
            "dev_105_vs_dhis2": dev_105_vs_dhis2
        })
    
    session = create_session(db, session_dict, lines_list)
    return session

@app.get("/export")
def export_csv(db: Session = Depends(get_db)):
    """Export all DQA lines as Excel with color-coded percentage deviations"""
    lines = db.query(DqaLine).join(DqaSession).join(Facility).join(Indicator).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DQA Data"
    
    # Define colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Header row
    headers = [
        "district", "facility", "period", "indicator_code", "indicator_name",
        "recount_register", "figure_105", "figure_dhis2",
        "dev_dhis2_vs_reg", "dev_105_vs_reg", "dev_105_vs_dhis2"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Deviation column indices (1-based, so 9, 10, 11)
    dev_cols = [9, 10, 11]
    
    # Data rows
    for row_idx, line in enumerate(lines, start=2):
        row_data = [
            line.session.facility.district,
            line.session.facility.name,
            line.session.period,
            line.indicator.code,
            line.indicator.name,
            line.recount_register,
            line.figure_105,
            line.figure_dhis2,
            line.dev_dhis2_vs_reg,
            line.dev_105_vs_reg,
            line.dev_105_vs_dhis2
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format deviation columns as percentages and color code
            if col_idx in dev_cols:
                if value is not None:
                    # Set as percentage
                    cell.value = value
                    cell.number_format = '0.0%'
                    
                    # Color code based on absolute deviation
                    abs_dev = abs(value)
                    if abs_dev <= 0.05:  # ≤ 5%
                        cell.fill = green_fill
                    elif abs_dev <= 0.10:  # 5-10%
                        cell.fill = yellow_fill
                    else:  # > 10%
                        cell.fill = red_fill
                else:
                    # Empty cell for null values
                    cell.value = ""
                    cell.fill = gray_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
    
    # Add comments section at the bottom
    # Group lines by session to add comments per facility
    sessions_dict = {}
    for line in lines:
        session_id = line.session.id
        if session_id not in sessions_dict:
            sessions_dict[session_id] = {
                "facility": line.session.facility.name,
                "district": line.session.facility.district,
                "period": line.session.period,
                "comments": line.session.comments
            }
    
    # Add comments rows after data
    if sessions_dict:
        comment_start_row = len(lines) + 3  # 2 rows gap after data
        ws.cell(row=comment_start_row, column=1, value="COMMENTS").font = Font(bold=True, size=12)
        comment_row = comment_start_row + 1
        
        for session_id, session_info in sessions_dict.items():
            if session_info["comments"]:
                ws.cell(row=comment_row, column=1, value=f"{session_info['facility']} ({session_info['district']}) - {session_info['period']}:").font = Font(bold=True)
                ws.cell(row=comment_row, column=2, value=session_info["comments"])
                # Merge cells for comment text (columns 2-11)
                ws.merge_cells(start_row=comment_row, start_column=2, end_row=comment_row, end_column=11)
                comment_row += 1
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dqa_export.xlsx"}
    )


@app.get("/export/session/{session_id}")
def export_session_csv(session_id: int, db: Session = Depends(get_db)):
    """Export a single DQA session as Excel with color-coded percentage deviations"""
    session = db.query(DqaSession).filter(DqaSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    lines = db.query(DqaLine).join(Indicator).filter(DqaLine.session_id == session_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "DQA Data"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    headers = [
        "district", "facility", "period", "indicator_code", "indicator_name",
        "recount_register", "figure_105", "figure_dhis2",
        "dev_dhis2_vs_reg", "dev_105_vs_reg", "dev_105_vs_dhis2"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    dev_cols = [9, 10, 11]

    for row_idx, line in enumerate(lines, start=2):
        row_data = [
            session.facility.district,
            session.facility.name,
            session.period,
            line.indicator.code,
            line.indicator.name,
            line.recount_register,
            line.figure_105,
            line.figure_dhis2,
            line.dev_dhis2_vs_reg,
            line.dev_105_vs_reg,
            line.dev_105_vs_dhis2,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_idx in dev_cols:
                if value is not None:
                    cell.value = value
                    cell.number_format = "0.0%"

                    abs_dev = abs(value)
                    if abs_dev <= 0.05:
                        cell.fill = green_fill
                    elif abs_dev <= 0.10:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
                else:
                    cell.value = ""
                    cell.fill = gray_fill

    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15

    # Add comments section at the bottom if comments exist
    if session.comments:
        comment_start_row = len(lines) + 3  # 2 rows gap after data
        ws.cell(row=comment_start_row, column=1, value="COMMENTS").font = Font(bold=True, size=12)
        comment_row = comment_start_row + 1
        ws.cell(row=comment_row, column=1, value=f"{session.facility.name} ({session.facility.district}) - {session.period}:").font = Font(bold=True)
        ws.cell(row=comment_row, column=2, value=session.comments)
        # Merge cells for comment text (columns 2-11)
        ws.merge_cells(start_row=comment_row, start_column=2, end_row=comment_row, end_column=11)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"dqa_session_{session.id}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/sessions/upload-csv")
def upload_csv(
    file: UploadFile = File(...),
    team: str = Form(None),
    db: Session = Depends(get_db)
):
    """Upload CSV file and create session(s)
    
    Args:
        file: CSV file to upload
        team: Team name to assign the uploaded data to (optional, defaults to 'csv_upload')
    """
    content = file.file.read().decode('utf-8')
    csv_rows = list(csv.DictReader(io.StringIO(content)))
    
    # First pass: collect comments from COMMENTS rows
    comments_by_session = {}
    for row in csv_rows:
        indicator_name = row.get("indicator_name", "").strip()
        if indicator_name.upper() == "COMMENTS":
            facility_name = row.get("facility", "").strip()
            district = row.get("district", "").strip()
            period = row.get("period", "").strip()
            comment_text = row.get("recount_register", "").strip() or None
            key = (facility_name, district, period)
            comments_by_session[key] = comment_text
    
    # Second pass: process indicator rows
    sessions_dict = {}
    for row in csv_rows:
        facility_name = row.get("facility", "").strip()
        district = row.get("district", "").strip()
        period = row.get("period", "").strip()
        indicator_code = row.get("indicator_code", "").strip()
        indicator_name = row.get("indicator_name", "").strip()
        
        # Skip COMMENTS rows in second pass
        if indicator_name.upper() == "COMMENTS":
            continue
        
        # Find facility
        facility = db.query(Facility).filter(
            Facility.name == facility_name,
            Facility.district == district
        ).first()
        
        if not facility:
            raise HTTPException(
                status_code=400,
                detail=f"Facility '{facility_name}' in district '{district}' not found"
            )
        
        # Find indicator - support both indicator_code and indicator_name
        indicator = None
        if indicator_code:
            indicator = get_indicator_by_code(db, indicator_code)
        elif indicator_name:
            indicator = db.query(Indicator).filter(Indicator.name == indicator_name).first()
        
        if not indicator:
            raise HTTPException(
                status_code=400,
                detail=f"Indicator not found. Please provide either 'indicator_code' or 'indicator_name'"
            )
        
        # Group key
        key = (facility.id, period)
        
        if key not in sessions_dict:
            # Use the team parameter from the form, or from CSV row, or default
            assigned_team = team or row.get("team", "").strip() or "csv_upload"
            # Get comments from comments_by_session if available
            comment_key = (facility_name, district, period)
            session_comments = comments_by_session.get(comment_key)
            sessions_dict[key] = {
                "facility_id": facility.id,
                "period": period,
                "team": assigned_team,
                "comments": session_comments,
                "lines": []
            }
        
        # Parse values
        recount_register = float(row.get("recount_register", "")) if row.get("recount_register", "").strip() else None
        figure_105 = float(row.get("figure_105", "")) if row.get("figure_105", "").strip() else None
        figure_dhis2 = float(row.get("figure_dhis2", "")) if row.get("figure_dhis2", "").strip() else None
        
        # Calculate deviations
        dev_dhis2_vs_reg, dev_105_vs_reg, dev_105_vs_dhis2 = calculate_deviations(
            recount_register, figure_105, figure_dhis2
        )
        
        sessions_dict[key]["lines"].append({
            "indicator_id": indicator.id,
            "recount_register": recount_register,
            "figure_105": figure_105,
            "figure_dhis2": figure_dhis2,
            "dev_dhis2_vs_reg": dev_dhis2_vs_reg,
            "dev_105_vs_reg": dev_105_vs_reg,
            "dev_105_vs_dhis2": dev_105_vs_dhis2
        })
    
    # Create sessions
    created_sessions = []
    for session_data in sessions_dict.values():
        session = create_session(db, {
            "facility_id": session_data["facility_id"],
            "period": session_data["period"],
            "team": session_data["team"],
            "comments": session_data.get("comments")
        }, session_data["lines"])
        created_sessions.append(session)
    
    return {"message": f"Created {len(created_sessions)} session(s)", "sessions": [s.id for s in created_sessions]}

@app.get("/reports/enhanced")
def generate_enhanced_report(db: Session = Depends(get_db)):
    """Generate enhanced Excel report with charts, summaries, and multi-sheet structure"""
    # Get all data
    lines = db.query(DqaLine).join(DqaSession).join(Facility).join(Indicator).all()
    stats = get_dashboard_stats(db)
    sessions = get_sessions(db)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # ========== SHEET 1: Executive Summary ==========
    ws_summary = wb.create_sheet("Executive Summary")
    
    # Title
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = "DQA Assessment - Executive Summary"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # Key Metrics Section
    row = 3
    ws_summary[f'A{row}'] = "Key Metrics"
    ws_summary[f'A{row}'].font = title_font
    
    row += 1
    metrics = [
        ["Total Facilities", stats['total_facilities']],
        ["Assessed Facilities", stats['assessed_facilities']],
        ["Completion Rate", f"{round((stats['assessed_facilities'] / stats['total_facilities'] * 100) if stats['total_facilities'] > 0 else 0)}%"],
        ["Total Sessions", stats['total_sessions']]
    ]
    
    for metric_name, metric_value in metrics:
        ws_summary[f'A{row}'] = metric_name
        ws_summary[f'B{row}'] = metric_value
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Team Performance Summary
    row += 2
    ws_summary[f'A{row}'] = "Team Performance Summary"
    ws_summary[f'A{row}'].font = title_font
    
    row += 1
    ws_summary[f'A{row}'] = "Team"
    ws_summary[f'B{row}'] = "Facilities Assessed"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'B{row}'].font = header_font
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary[f'B{row}'].fill = header_fill
    
    for team_data in stats['team_progress']:
        row += 1
        ws_summary[f'A{row}'] = team_data['team']
        ws_summary[f'B{row}'] = team_data['facilities_assessed']
    
    # AI-Generated Executive Summary
    row += 3
    ws_summary[f'A{row}'] = "AI-Generated Executive Summary"
    ws_summary[f'A{row}'].font = title_font
    
    # Generate AI summary
    ai_summary = generate_ai_summary(stats, sessions, lines)
    
    if ai_summary:
        row += 1
        # Split summary into paragraphs and add to cells
        paragraphs = ai_summary.split('\n\n')
        for para in paragraphs:
            if para.strip():
                ws_summary[f'A{row}'] = para.strip()
                ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                # Merge cells for better readability (columns A-D)
                ws_summary.merge_cells(f'A{row}:D{row}')
                row += 1
    else:
        row += 1
        ws_summary[f'A{row}'] = "AI summary not available. Set GEMINI_API_KEY environment variable to enable AI-generated insights."
        ws_summary[f'A{row}'].font = Font(italic=True, color="808080")
        ws_summary.merge_cells(f'A{row}:D{row}')
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 30
    ws_summary.column_dimensions['D'].width = 30
    
    # ========== SHEET 2: Charts & Visualizations ==========
    ws_charts = wb.create_sheet("Charts & Visualizations")
    
    # Overall Progress Pie Chart Data
    ws_charts['A1'] = "Overall Progress"
    ws_charts['A1'].font = title_font
    ws_charts['A3'] = "Category"
    ws_charts['B3'] = "Count"
    ws_charts['A3'].font = header_font
    ws_charts['B3'].font = header_font
    ws_charts['A3'].fill = header_fill
    ws_charts['B3'].fill = header_fill
    
    ws_charts['A4'] = "Assessed"
    ws_charts['B4'] = stats['assessed_facilities']
    ws_charts['A5'] = "Remaining"
    ws_charts['B5'] = stats['total_facilities'] - stats['assessed_facilities']
    
    # Create Pie Chart
    pie = PieChart()
    pie.title = "Overall Progress"
    data = Reference(ws_charts, min_col=2, min_row=3, max_row=5)
    cats = Reference(ws_charts, min_col=1, min_row=4, max_row=5)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(cats)
    pie.width = 10
    pie.height = 7
    ws_charts.add_chart(pie, "D2")
    
    # Team Progress Bar Chart Data
    chart_row = 10
    ws_charts[f'A{chart_row}'] = "Team Progress"
    ws_charts[f'A{chart_row}'].font = title_font
    chart_row += 1
    ws_charts[f'A{chart_row}'] = "Team"
    ws_charts[f'B{chart_row}'] = "Facilities Assessed"
    ws_charts[f'A{chart_row}'].font = header_font
    ws_charts[f'B{chart_row}'].font = header_font
    ws_charts[f'A{chart_row}'].fill = header_fill
    ws_charts[f'B{chart_row}'].fill = header_fill
    
    chart_row += 1
    for team_data in stats['team_progress']:
        ws_charts[f'A{chart_row}'] = team_data['team']
        ws_charts[f'B{chart_row}'] = team_data['facilities_assessed']
        chart_row += 1
    
    # Create Bar Chart
    bar = BarChart()
    bar.title = "Facilities Assessed by Team"
    bar.type = "col"
    bar.style = 10
    data = Reference(ws_charts, min_col=2, min_row=chart_row - len(stats['team_progress']), max_row=chart_row - 1)
    cats = Reference(ws_charts, min_col=1, min_row=chart_row - len(stats['team_progress']), max_row=chart_row - 1)
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(cats)
    bar.width = 10
    bar.height = 7
    ws_charts.add_chart(bar, "D10")
    
    ws_charts.column_dimensions['A'].width = 20
    ws_charts.column_dimensions['B'].width = 20
    
    # ========== SHEET 3: Detailed Data (existing format) ==========
    ws_data = wb.create_sheet("Detailed Data")
    
    # Header row
    headers = [
        "district", "facility", "period", "indicator_code", "indicator_name",
        "recount_register", "figure_105", "figure_dhis2",
        "dev_dhis2_vs_reg", "dev_105_vs_reg", "dev_105_vs_dhis2"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.font = header_font
    
    # Deviation column indices
    dev_cols = [9, 10, 11]
    
    # Data rows
    for row_idx, line in enumerate(lines, start=2):
        row_data = [
            line.session.facility.district,
            line.session.facility.name,
            line.session.period,
            line.indicator.code,
            line.indicator.name,
            line.recount_register,
            line.figure_105,
            line.figure_dhis2,
            line.dev_dhis2_vs_reg,
            line.dev_105_vs_reg,
            line.dev_105_vs_dhis2
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            
            if col_idx in dev_cols:
                if value is not None:
                    cell.value = value
                    cell.number_format = '0.0%'
                    abs_dev = abs(value)
                    if abs_dev <= 0.05:
                        cell.fill = green_fill
                    elif abs_dev <= 0.10:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
                else:
                    cell.value = ""
                    cell.fill = gray_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws_data.column_dimensions[column_letter].width = 15
    
    # Add comments section at the bottom
    sessions_dict = {}
    for line in lines:
        session_id = line.session.id
        if session_id not in sessions_dict:
            sessions_dict[session_id] = {
                "facility": line.session.facility.name,
                "district": line.session.facility.district,
                "period": line.session.period,
                "comments": line.session.comments
            }
    
    if sessions_dict:
        comment_start_row = len(lines) + 3
        ws_data.cell(row=comment_start_row, column=1, value="COMMENTS").font = Font(bold=True, size=12)
        comment_row = comment_start_row + 1
        
        for session_id, session_info in sessions_dict.items():
            if session_info["comments"]:
                ws_data.cell(row=comment_row, column=1, value=f"{session_info['facility']} ({session_info['district']}) - {session_info['period']}:").font = Font(bold=True)
                ws_data.cell(row=comment_row, column=2, value=session_info["comments"])
                ws_data.merge_cells(start_row=comment_row, start_column=2, end_row=comment_row, end_column=11)
                comment_row += 1
    
    # ========== SHEET 4: Team Analysis ==========
    ws_teams = wb.create_sheet("Team Analysis")
    
    ws_teams['A1'] = "Team Analysis"
    ws_teams['A1'].font = title_font
    
    row = 3
    ws_teams[f'A{row}'] = "Team"
    ws_teams[f'B{row}'] = "Facilities Assessed"
    ws_teams[f'C{row}'] = "Total Sessions"
    ws_teams[f'A{row}'].font = header_font
    ws_teams[f'B{row}'].font = header_font
    ws_teams[f'C{row}'].font = header_font
    ws_teams[f'A{row}'].fill = header_fill
    ws_teams[f'B{row}'].fill = header_fill
    ws_teams[f'C{row}'].fill = header_fill
    
    # Calculate sessions per team
    team_sessions = {}
    for session in sessions:
        team = session.get('team', 'Unknown')
        if team not in team_sessions:
            team_sessions[team] = 0
        team_sessions[team] += 1
    
    row += 1
    for team_data in stats['team_progress']:
        team_name = team_data['team']
        ws_teams[f'A{row}'] = team_name
        ws_teams[f'B{row}'] = team_data['facilities_assessed']
        ws_teams[f'C{row}'] = team_sessions.get(team_name, 0)
        row += 1
    
    # AI-Generated Team Insights
    row += 2
    ws_teams[f'A{row}'] = "AI-Generated Team Insights"
    ws_teams[f'A{row}'].font = title_font
    
    ai_insights = generate_ai_team_insights(stats, sessions)
    
    if ai_insights:
        row += 1
        # Split insights into lines and add to cells
        insight_lines = ai_insights.split('\n')
        for line in insight_lines:
            if line.strip():
                ws_teams[f'A{row}'] = line.strip()
                ws_teams[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws_teams.merge_cells(f'A{row}:C{row}')
                row += 1
    else:
        row += 1
        ws_teams[f'A{row}'] = "AI insights not available. Set GEMINI_API_KEY environment variable to enable AI-generated insights."
        ws_teams[f'A{row}'].font = Font(italic=True, color="808080")
        ws_teams.merge_cells(f'A{row}:C{row}')
    
    ws_teams.column_dimensions['A'].width = 20
    ws_teams.column_dimensions['B'].width = 20
    ws_teams.column_dimensions['C'].width = 20
    
    # ========== SHEET 5: Comments ==========
    ws_comments = wb.create_sheet("Comments")
    
    ws_comments['A1'] = "Facility Comments"
    ws_comments['A1'].font = title_font
    
    row = 3
    ws_comments[f'A{row}'] = "Facility"
    ws_comments[f'B{row}'] = "District"
    ws_comments[f'C{row}'] = "Period"
    ws_comments[f'D{row}'] = "Comments"
    ws_comments[f'A{row}'].font = header_font
    ws_comments[f'B{row}'].font = header_font
    ws_comments[f'C{row}'].font = header_font
    ws_comments[f'D{row}'].font = header_font
    ws_comments[f'A{row}'].fill = header_fill
    ws_comments[f'B{row}'].fill = header_fill
    ws_comments[f'C{row}'].fill = header_fill
    ws_comments[f'D{row}'].fill = header_fill
    
    row += 1
    for session_id, session_info in sessions_dict.items():
        if session_info["comments"]:
            ws_comments[f'A{row}'] = session_info['facility']
            ws_comments[f'B{row}'] = session_info['district']
            ws_comments[f'C{row}'] = session_info['period']
            ws_comments[f'D{row}'] = session_info['comments']
            ws_comments[f'D{row}'].alignment = Alignment(wrap_text=True)
            row += 1
    
    if row == 4:  # No comments found
        ws_comments['A4'] = "No comments available"
    
    ws_comments.column_dimensions['A'].width = 30
    ws_comments.column_dimensions['B'].width = 20
    ws_comments.column_dimensions['C'].width = 20
    ws_comments.column_dimensions['D'].width = 50
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dqa_enhanced_report.xlsx"}
    )

